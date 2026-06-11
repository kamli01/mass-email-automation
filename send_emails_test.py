"""
send_certificates_test.py
-------------------------
TEST script -- uses participants2.xlsx
Matches each participant's name against certificate PNG files in this folder,
then sends the matched certificate to their email address.

HOW TO USE
----------
1. Fill in your Gmail credentials in the CONFIG section below.
2. Run once with DRY_RUN = True to verify the name <-> file matching.
3. Set DRY_RUN = False to actually send the emails.

REQUIREMENTS
------------
    pip install openpyxl

GMAIL SETUP (App Password)
--------------------------
If you have 2-Step Verification on your Google account (recommended), you
must create an App Password instead of using your normal password:
  Google Account -> Security -> 2-Step Verification -> App Passwords
  Create an app password, copy the 16-character code, paste it below.
"""

import os
import smtplib
import openpyxl
import uuid
from datetime import datetime, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from difflib import get_close_matches

# --------------------------- CONFIG ------------------------------------------
SENDER_NAME     = "GDGoC TCOER"                # Display name shown in inbox
SENDER_EMAIL    = "dsctcoer2021@gmail.com"
SENDER_PASSWORD = "zbpz wnet aqmv egyz"        # Gmail App Password
SMTP_HOST       = "smtp.gmail.com"
SMTP_PORT       = 587

XLSX_FILE       = "participants2.xlsx"         # test file
CERTS_DIR       = "."                          # folder containing certificate PNGs

# Set to True to preview matches without actually sending emails
DRY_RUN         = False

# Email subject (use {name} as placeholder)
EMAIL_SUBJECT   = "Your Certificate of Participation - GDGoC TCOER Workshop"

# Plain-text fallback (for email clients that don't render HTML)
EMAIL_BODY_TEXT = """\
Hi {name},

Congratulations on completing the "Build with AI on Flutter" Workshop organized by GDGoC TCOER!

Please find your Certificate of Participation attached to this email.

We look forward to seeing you at our future events. Stay connected with GDGoC TCOER!

Best regards,
GDGoC TCOER Team
dsctcoer2021@gmail.com
"""

# HTML version (renders nicely, helps avoid spam)
EMAIL_BODY_HTML = """\
<!DOCTYPE html>
<html>
<body style="margin:0;padding:0;background-color:#f4f4f4;font-family:Arial,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background-color:#f4f4f4;padding:30px 0;">
    <tr>
      <td align="center">
        <table width="600" cellpadding="0" cellspacing="0"
               style="background-color:#ffffff;border-radius:8px;overflow:hidden;
                      box-shadow:0 2px 8px rgba(0,0,0,0.08);">

          <!-- Header -->
          <tr>
            <td style="background-color:#1a73e8;padding:28px 40px;">
              <h1 style="margin:0;color:#ffffff;font-size:22px;font-weight:700;">
                GDGoC TCOER
              </h1>
              <p style="margin:4px 0 0;color:#c8dcfa;font-size:13px;">
                Google Developer Group on Campus
              </p>
            </td>
          </tr>

          <!-- Body -->
          <tr>
            <td style="padding:36px 40px;">
              <p style="margin:0 0 16px;font-size:16px;color:#202124;">
                Hi <strong>{name}</strong>,
              </p>
              <p style="margin:0 0 16px;font-size:15px;color:#3c4043;line-height:1.6;">
                Congratulations on completing the
                <strong>"Build with AI on Flutter"</strong> Workshop
                organized by <strong>GDGoC TCOER</strong>!
              </p>
              <p style="margin:0 0 24px;font-size:15px;color:#3c4043;line-height:1.6;">
                Please find your <strong>Certificate of Participation</strong>
                attached to this email.
              </p>

              <!-- CTA box -->
              <table cellpadding="0" cellspacing="0" width="100%"
                     style="background-color:#f0f7ff;border-left:4px solid #1a73e8;
                            border-radius:4px;margin-bottom:24px;">
                <tr>
                  <td style="padding:16px 20px;font-size:14px;color:#1a73e8;">
                    We look forward to seeing you at our future events.
                    Stay connected with GDGoC TCOER!
                  </td>
                </tr>
              </table>

              <p style="margin:0;font-size:14px;color:#5f6368;">
                Best regards,<br>
                <strong>GDGoC TCOER Team</strong>
              </p>
            </td>
          </tr>

          <!-- Footer -->
          <tr>
            <td style="background-color:#f8f9fa;padding:16px 40px;
                       border-top:1px solid #e8eaed;">
              <p style="margin:0;font-size:12px;color:#80868b;text-align:center;">
                This email was sent by GDGoC TCOER &bull;
                <a href="mailto:dsctcoer2021@gmail.com"
                   style="color:#1a73e8;text-decoration:none;">
                  dsctcoer2021@gmail.com
                </a>
              </p>
            </td>
          </tr>

        </table>
      </td>
    </tr>
  </table>
</body>
</html>
"""
# -----------------------------------------------------------------------------


def load_participants(xlsx_path: str):
    """Return list of (name, email) tuples from the xlsx file."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    name_col = email_col = None
    for i, h in enumerate(headers):
        if "name" in h:
            name_col = i
        if "email" in h or "e-mail" in h:
            email_col = i

    if name_col is None or email_col is None:
        raise ValueError(f"Could not detect name/email columns. Headers: {headers}")

    participants = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name  = str(row[name_col]).strip()  if row[name_col]  else ""
        email = str(row[email_col]).strip() if row[email_col] else ""
        if name and email and name != "None" and email != "None":
            participants.append((name, email))
    return participants


def get_certificate_files(certs_dir: str):
    """Return a dict mapping lower-cased base-name -> full file path for all *_certificate.png files."""
    cert_map = {}
    for fname in os.listdir(certs_dir):
        if fname.lower().endswith("_certificate.png"):
            base = fname[: fname.lower().rfind("_certificate.png")].lower()
            cert_map[base] = os.path.join(certs_dir, fname)
    return cert_map


def find_certificate(name: str, cert_map: dict):
    """Find certificate file for a given name (exact then fuzzy match)."""
    name_lower = name.lower()
    if name_lower in cert_map:
        return cert_map[name_lower]
    candidates = list(cert_map.keys())
    matches = get_close_matches(name_lower, candidates, n=1, cutoff=0.7)
    if matches:
        return cert_map[matches[0]]
    return None


def send_email(sender_name: str, sender: str, password: str, recipient: str,
               subject: str, body_text: str, body_html: str, attachment_path: str):
    """Send a multipart HTML+text email with an attachment via Gmail SMTP."""

    # multipart/mixed holds the message + attachment
    msg = MIMEMultipart("mixed")
    msg["From"]       = f"{sender_name} <{sender}>"
    msg["To"]         = recipient
    msg["Subject"]    = subject
    msg["Reply-To"]   = f"{sender_name} <{sender}>"
    msg["Date"]       = datetime.now(timezone.utc).strftime("%a, %d %b %Y %H:%M:%S +0000")
    msg["Message-ID"] = f"<{uuid.uuid4()}@{sender.split('@')[1]}>"

    # Inline body: plain + HTML alternative
    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(body_text, "plain", "utf-8"))
    alt.attach(MIMEText(body_html, "html",  "utf-8"))
    msg.attach(alt)

    # Attachment
    with open(attachment_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header(
        "Content-Disposition",
        f'attachment; filename="{os.path.basename(attachment_path)}"',
    )
    msg.attach(part)

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.ehlo()
        server.starttls()
        server.login(sender, password.replace(" ", ""))
        server.sendmail(sender, recipient, msg.as_string())


def main():
    print(f"{'[DRY RUN] ' if DRY_RUN else ''}Loading participants from: {XLSX_FILE}\n")
    participants = load_participants(XLSX_FILE)
    cert_map     = get_certificate_files(CERTS_DIR)

    print(f"Found {len(participants)} participant(s) and {len(cert_map)} certificate(s).\n")
    print(f"{'Name':<35} {'Email':<40} {'Certificate Found?'}")
    print("-" * 100)

    success_count = 0
    skipped       = []

    for name, email in participants:
        cert_path = find_certificate(name, cert_map)
        status    = f"[OK] {os.path.basename(cert_path)}" if cert_path else "[!!] NOT FOUND"
        print(f"{name:<35} {email:<40} {status}")

        if cert_path:
            if not DRY_RUN:
                try:
                    send_email(
                        sender_name     = SENDER_NAME,
                        sender          = SENDER_EMAIL,
                        password        = SENDER_PASSWORD,
                        recipient       = email,
                        subject         = EMAIL_SUBJECT.format(name=name),
                        body_text       = EMAIL_BODY_TEXT.format(name=name),
                        body_html       = EMAIL_BODY_HTML.format(name=name),
                        attachment_path = cert_path,
                    )
                    print(f"   -> Email sent to {email}")
                    success_count += 1
                except Exception as e:
                    print(f"   -> FAILED to send to {email}: {e}")
                    skipped.append((name, email, str(e)))
            else:
                success_count += 1
        else:
            skipped.append((name, email, "No matching certificate file"))

    print("\n" + "=" * 100)
    if DRY_RUN:
        print(f"[DRY RUN] Would send {success_count} email(s). Set DRY_RUN = False to actually send.")
    else:
        print(f"Sent: {success_count}  |  Skipped/Failed: {len(skipped)}")

    if skipped:
        print("\nSkipped / Failed:")
        for entry in skipped:
            print(f"  * {entry[0]} <{entry[1]}> -- {entry[2]}")


if __name__ == "__main__":
    main()
